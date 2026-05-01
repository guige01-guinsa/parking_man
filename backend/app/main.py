from __future__ import annotations

import base64
from io import BytesIO
import json
import os
import re
import shutil
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from .auth import COOKIE_NAME, make_session, pbkdf2_hash, pbkdf2_verify, read_session, require_role
from .db import DEFAULT_SITE_CODE, DEFAULT_SITE_NAME, connect, init_db, maybe_seed_demo, normalize_site_code, seed_users
from .excel_import import describe_excel_files, store_registry_upload, sync_registry_from_dir
from .ocr_learning import get_learning_candidates, get_learning_status, parse_candidates_json, record_ocr_feedback
from .ocr import scan_plate_image
from .plates import PlateVerdict, evaluate_vehicle_row, normalize_plate

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
UPLOAD_DIR = Path(os.getenv("PARKING_UPLOAD_DIR", str(BASE_DIR / "uploads")))
IMPORT_DIR = Path(os.getenv("PARKING_IMPORT_DIR", str(BASE_DIR.parent / "imports")))
APP_TITLE = os.getenv("PARKING_APP_TITLE", "아파트 주차단속 시스템")
ROOT_PATH = os.getenv("PARKING_ROOT_PATH", "").strip()
LOCAL_LOGIN_ENABLED = os.getenv("PARKING_LOCAL_LOGIN_ENABLED", "1").strip().lower() in {"1", "true", "yes", "on"}
SUPPORT_KAKAO_URL = os.getenv("PARKING_SUPPORT_KAKAO_URL", "").strip()
SUPPORT_KAKAO_LABEL = os.getenv("PARKING_SUPPORT_KAKAO_LABEL", "카카오톡 문의").strip() or "카카오톡 문의"
BILLING_PROVIDER = os.getenv("PARKING_BILLING_PROVIDER", "manual").strip().lower() or "manual"
BILLING_ENFORCEMENT_ENABLED = os.getenv("PARKING_BILLING_ENFORCEMENT_ENABLED", "0").strip().lower() in {"1", "true", "yes", "on"}
SALES_CONTACT_URL = os.getenv("PARKING_SALES_CONTACT_URL", "").strip()
GOOGLE_PLAY_PACKAGE_NAME = os.getenv("PARKING_GOOGLE_PLAY_PACKAGE_NAME", "com.parkingmanagement.app").strip() or "com.parkingmanagement.app"
GOOGLE_PLAY_SERVICE_ACCOUNT_JSON = os.getenv("PARKING_GOOGLE_PLAY_SERVICE_ACCOUNT_JSON", "").strip()
GOOGLE_PLAY_SERVICE_ACCOUNT_FILE = os.getenv("PARKING_GOOGLE_PLAY_SERVICE_ACCOUNT_FILE", os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "")).strip()
GOOGLE_PLAY_AUTO_ACKNOWLEDGE = os.getenv("PARKING_GOOGLE_PLAY_AUTO_ACKNOWLEDGE", "1").strip().lower() in {"1", "true", "yes", "on"}
GOOGLE_PLAY_RTDN_TOKEN = os.getenv("PARKING_GOOGLE_PLAY_RTDN_TOKEN", "").strip()
GOOGLE_PLAY_PRODUCT_IDS = {
    "starter": os.getenv("PARKING_GOOGLE_PLAY_PRODUCT_STARTER", "parking_starter_monthly").strip() or "parking_starter_monthly",
    "standard": os.getenv("PARKING_GOOGLE_PLAY_PRODUCT_STANDARD", "parking_standard_monthly").strip() or "parking_standard_monthly",
    "pro": os.getenv("PARKING_GOOGLE_PLAY_PRODUCT_PRO", "parking_pro_monthly").strip() or "parking_pro_monthly",
}

if ROOT_PATH and not ROOT_PATH.startswith("/"):
    ROOT_PATH = f"/{ROOT_PATH}"
ROOT_PATH = ROOT_PATH.rstrip("/")

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
IMPORT_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title=APP_TITLE, version="2.0.0", root_path=ROOT_PATH)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
_ready_lock = threading.Lock()
_app_ready = False
LOGIN_INVALID_MESSAGE = "로그인에 실패했습니다. 아이디와 비밀번호를 다시 확인해 주세요."
LOGIN_FORMAT_MESSAGE = "아이디 형식을 다시 확인해 주세요. 영문 소문자와 숫자를 사용해 입력할 수 있습니다."
LOGIN_SITE_FORMAT_MESSAGE = "아파트 코드는 영문, 숫자, 하이픈(-), 밑줄(_)만 사용해 2~32자로 입력해 주세요."


class CheckMatch(BaseModel):
    plate: str
    verdict: str
    message: str
    unit: str | None = None
    building: str | None = None
    unit_number: str | None = None
    owner_name: str | None = None
    phone: str | None = None
    status: str | None = None
    valid_from: str | None = None
    valid_to: str | None = None


class CheckResponse(CheckMatch):
    site_code: str
    requested_plate: str | None = None
    match_mode: str = "exact"
    match_count: int = 1
    match_index: int = 0
    matches: list[CheckMatch] = Field(default_factory=list)


class UserCreateRequest(BaseModel):
    username: str
    password: str
    role: str
    can_manage_vehicles: bool = False


class UserUpdateRequest(BaseModel):
    role: str | None = None
    password: str | None = None
    can_manage_vehicles: bool | None = None


class VehicleUpsertRequest(BaseModel):
    plate: str
    unit: str | None = None
    building: str | None = None
    unit_number: str | None = None
    owner_name: str | None = None
    phone: str | None = None
    status: str | None = "active"
    valid_from: str | None = None
    valid_to: str | None = None
    note: str | None = None


class RegistrySyncRequest(BaseModel):
    preserve_manual: bool = True


class SiteCreateRequest(BaseModel):
    site_code: str
    name: str
    admin_username: str
    admin_password: str


class BillingInquiryRequest(BaseModel):
    requested_plan: str
    contact_name: str | None = None
    contact_phone: str | None = None
    contact_email: str | None = None
    message: str | None = None


class GooglePlayVerifyRequest(BaseModel):
    product_id: str
    purchase_token: str
    package_name: str | None = None


class CctvAssignmentRequest(BaseModel):
    location: str | None = None
    search_start_time: str | None = None
    search_end_time: str | None = None
    content: str | None = None
    assigned_to: str | None = None
    work_weight: int | None = Field(None, ge=1, le=5)
    instruction: str | None = None
    status: str | None = None


class EnforcementEventUpdateRequest(BaseModel):
    plate: str | None = None
    inspector: str | None = None
    location: str | None = None
    memo: str | None = None


USERNAME_PATTERN = re.compile(r"^[a-z0-9][a-z0-9._-]{2,31}$")
SITE_CODE_PATTERN = re.compile(r"^[A-Z0-9][A-Z0-9_-]{1,31}$")
ROLE_LABELS = {
    "admin": "관리자",
    "director": "소장",
    "manager": "과장",
    "section_chief": "계장",
    "team_lead": "팀장",
    "staff": "주임",
    "guard": "경비",
    "cleaner": "미화",
}
VALID_USER_ROLES = set(ROLE_LABELS)
ROLE_ORDER = list(ROLE_LABELS)
VIEW_ROLES = set(ROLE_LABELS)
ENFORCEMENT_WRITE_ROLES = {"admin", "director", "manager", "section_chief", "team_lead", "staff", "guard"}
CCTV_ASSIGNMENT_ROLES = {"admin", "director", "manager", "section_chief", "team_lead"}
CCTV_STATUS_LABELS = {
    "requested": "요청",
    "assigned": "배정",
    "in_progress": "진행",
    "done": "완료",
    "cancelled": "취소",
}
CCTV_STATUSES = set(CCTV_STATUS_LABELS)
BILLING_PLAN_CATALOG = {
    "trial": {
        "code": "trial",
        "name": "무료 체험",
        "monthly_price_krw": 0,
        "users_limit": 3,
        "vehicles_limit": 300,
        "monthly_records_limit": 1000,
        "monthly_cctv_limit": 20,
        "support": "체험 지원",
    },
    "starter": {
        "code": "starter",
        "name": "Starter",
        "monthly_price_krw": 49000,
        "users_limit": 5,
        "vehicles_limit": 1000,
        "monthly_records_limit": 3000,
        "monthly_cctv_limit": 100,
        "support": "이메일 지원",
    },
    "standard": {
        "code": "standard",
        "name": "Standard",
        "monthly_price_krw": 99000,
        "users_limit": 15,
        "vehicles_limit": 5000,
        "monthly_records_limit": 15000,
        "monthly_cctv_limit": 500,
        "support": "우선 지원",
    },
    "pro": {
        "code": "pro",
        "name": "Pro",
        "monthly_price_krw": 199000,
        "users_limit": 50,
        "vehicles_limit": 20000,
        "monthly_records_limit": 50000,
        "monthly_cctv_limit": 2000,
        "support": "전담 지원",
    },
}
BILLING_STATUS_LABELS = {
    "trialing": "체험 중",
    "active": "사용 중",
    "past_due": "결제 필요",
    "cancelled": "해지",
}


def app_url(path: str) -> str:
    if not path.startswith("/"):
        path = f"/{path}"
    return f"{ROOT_PATH}{path}" if ROOT_PATH else path


def role_order_case(column: str = "role") -> str:
    clauses = " ".join(f"WHEN '{role}' THEN {index}" for index, role in enumerate(ROLE_ORDER))
    return f"CASE {column} {clauses} ELSE {len(ROLE_ORDER)} END"


def session_cookie_path() -> str:
    return ROOT_PATH or "/"


def render_login_page(
    request: Request,
    *,
    status_code: int = 200,
    username: str = "",
    site_code: str = DEFAULT_SITE_CODE,
    error: str | None = None,
):
    return templates.TemplateResponse(
        request=request,
        name="login.html",
        context={
            "app_title": APP_TITLE,
            "username": username,
            "site_code": site_code,
            "login_error": error,
            "support_kakao_url": SUPPORT_KAKAO_URL,
            "support_kakao_label": SUPPORT_KAKAO_LABEL,
        },
        status_code=status_code,
    )


def normalize_login_site_code(value: str | None) -> str:
    site_code = normalize_site_code(value)
    if not SITE_CODE_PATTERN.fullmatch(site_code):
        raise HTTPException(status_code=400, detail=LOGIN_SITE_FORMAT_MESSAGE)
    return site_code


def normalize_required_site_code(value: str | None) -> str:
    if not str(value or "").strip():
        raise HTTPException(status_code=400, detail="아파트 코드를 입력해 주세요.")
    return normalize_login_site_code(value)


def normalize_site_name(value: str | None) -> str:
    name = str(value or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="아파트명을 입력해 주세요.")
    if len(name) > 80:
        raise HTTPException(status_code=400, detail="아파트명은 80자 이내로 입력해 주세요.")
    return name


def site_storage_key(site_code: str) -> str:
    key = re.sub(r"[^A-Z0-9_-]+", "-", normalize_site_code(site_code)).strip("-_").lower()
    return key or normalize_site_code(DEFAULT_SITE_CODE).lower()


def site_import_dir(site_code: str) -> Path:
    normalized_site = normalize_site_code(site_code)
    if normalized_site == normalize_site_code(DEFAULT_SITE_CODE):
        return IMPORT_DIR
    return IMPORT_DIR / site_storage_key(normalized_site)


def site_upload_dir(site_code: str) -> Path:
    return UPLOAD_DIR / site_storage_key(site_code)


def site_upload_url(site_code: str, filename: str) -> str:
    return app_url(f"/uploads/{site_storage_key(site_code)}/{filename}")


def site_public_dict(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    return {
        "site_code": row["site_code"],
        "name": row["name"],
        "created_at": row["created_at"],
        "users_count": row.get("users_count", 0),
        "vehicles_count": row.get("vehicles_count", 0),
    }


def site_name_for_code(site_code: str) -> str:
    with connect() as con:
        row = con.execute("SELECT name FROM sites WHERE site_code = ?", (normalize_site_code(site_code),)).fetchone()
    return row["name"] if row else normalize_site_code(site_code)


def normalize_billing_plan(value: str | None) -> str:
    plan = str(value or "").strip().lower()
    if plan not in BILLING_PLAN_CATALOG or plan == "trial":
        raise HTTPException(status_code=400, detail="요금제를 다시 확인해 주세요.")
    return plan


def normalize_billing_text(value: str | None, field_name: str, max_length: int) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if len(text) > max_length:
        raise HTTPException(status_code=400, detail=f"{field_name}은 {max_length}자 이내로 입력해 주세요.")
    return text


def google_play_product_id_for_plan(plan_code: str) -> str | None:
    return GOOGLE_PLAY_PRODUCT_IDS.get(plan_code)


def plan_for_google_play_product(product_id: str) -> str:
    normalized = normalize_google_play_product_id(product_id)
    for plan_code, configured_product_id in GOOGLE_PLAY_PRODUCT_IDS.items():
        if configured_product_id == normalized:
            return plan_code
    raise HTTPException(status_code=400, detail="등록되지 않은 Google Play 상품입니다.")


def normalize_google_play_product_id(value: str | None) -> str:
    product_id = str(value or "").strip()
    if not product_id or len(product_id) > 200 or not re.fullmatch(r"[A-Za-z0-9._-]+", product_id):
        raise HTTPException(status_code=400, detail="Google Play 상품 ID를 다시 확인해 주세요.")
    return product_id


def normalize_purchase_token(value: str | None) -> str:
    token = str(value or "").strip()
    if not token or len(token) > 4096:
        raise HTTPException(status_code=400, detail="Google Play 구매 토큰을 다시 확인해 주세요.")
    return token


def google_play_configured() -> bool:
    return bool(GOOGLE_PLAY_PACKAGE_NAME and (GOOGLE_PLAY_SERVICE_ACCOUNT_JSON or GOOGLE_PLAY_SERVICE_ACCOUNT_FILE))


def billing_plan_public_dict(plan_code: str) -> dict[str, Any]:
    plan = dict(BILLING_PLAN_CATALOG[plan_code])
    plan["display_price"] = "무료" if int(plan["monthly_price_krw"]) <= 0 else f"월 {int(plan['monthly_price_krw']):,}원"
    plan["google_play_product_id"] = google_play_product_id_for_plan(plan_code)
    return plan


def ensure_site_billing_row(con, site_code: str) -> dict[str, Any]:
    normalized_site = normalize_site_code(site_code)
    con.execute(
        """
        INSERT OR IGNORE INTO site_billing(site_code, plan, status, trial_ends_at, payment_provider)
        VALUES (?, 'trial', 'trialing', date('now', '+14 days'), ?)
        """,
        (normalized_site, BILLING_PROVIDER),
    )
    row = con.execute("SELECT * FROM site_billing WHERE site_code = ?", (normalized_site,)).fetchone()
    return dict(row)


def billing_status_for_site(site_code: str) -> dict[str, Any]:
    normalized_site = normalize_site_code(site_code)
    with connect() as con:
        billing = ensure_site_billing_row(con, normalized_site)
        usage = {
            "users": con.execute("SELECT COUNT(*) AS cnt FROM users WHERE site_code = ?", (normalized_site,)).fetchone()["cnt"],
            "vehicles": con.execute("SELECT COUNT(*) AS cnt FROM vehicles WHERE site_code = ?", (normalized_site,)).fetchone()["cnt"],
            "monthly_records": con.execute(
                """
                SELECT COUNT(*) AS cnt
                FROM enforcement_events
                WHERE site_code = ? AND created_at >= datetime('now', 'start of month')
                """,
                (normalized_site,),
            ).fetchone()["cnt"],
            "monthly_cctv": con.execute(
                """
                SELECT COUNT(*) AS cnt
                FROM cctv_search_requests
                WHERE site_code = ? AND created_at >= datetime('now', 'start of month')
                """,
                (normalized_site,),
            ).fetchone()["cnt"],
        }
        latest_inquiries = [
            dict(row)
            for row in con.execute(
                """
                SELECT id, requested_plan, contact_name, contact_phone, contact_email, message, status, created_at
                FROM billing_inquiries
                WHERE site_code = ?
                ORDER BY id DESC
                LIMIT 5
                """,
                (normalized_site,),
            ).fetchall()
        ]
        latest_google_play_purchases = [
            dict(row)
            for row in con.execute(
                """
                SELECT id, product_id, plan, order_id, subscription_state, acknowledgement_state, expires_at, verified_at
                FROM google_play_purchases
                WHERE site_code = ?
                ORDER BY verified_at DESC, id DESC
                LIMIT 5
                """,
                (normalized_site,),
            ).fetchall()
        ]
        trial_days_remaining = con.execute(
            """
            SELECT
              CASE
                WHEN trial_ends_at IS NULL THEN NULL
                WHEN julianday(trial_ends_at) <= julianday('now') THEN 0
                ELSE CAST(julianday(trial_ends_at) - julianday('now') AS INTEGER) + 1
              END AS days
            FROM site_billing
            WHERE site_code = ?
            """,
            (normalized_site,),
        ).fetchone()["days"]
        con.commit()

    plan_code = billing.get("plan") if billing.get("plan") in BILLING_PLAN_CATALOG else "trial"
    plan = billing_plan_public_dict(plan_code)
    return {
        "site_code": normalized_site,
        "site_name": site_name_for_code(normalized_site),
        "billing": {
            "plan": plan_code,
            "plan_label": plan["name"],
            "status": billing.get("status") or "trialing",
            "status_label": BILLING_STATUS_LABELS.get(billing.get("status"), billing.get("status") or "체험 중"),
            "trial_ends_at": billing.get("trial_ends_at"),
            "trial_days_remaining": trial_days_remaining,
            "current_period_ends_at": billing.get("current_period_ends_at"),
            "payment_provider": billing.get("payment_provider") or BILLING_PROVIDER,
            "enforcement_enabled": BILLING_ENFORCEMENT_ENABLED,
        },
        "current_plan": plan,
        "plans": [billing_plan_public_dict(code) for code in ("starter", "standard", "pro")],
        "usage": usage,
        "latest_inquiries": latest_inquiries,
        "latest_google_play_purchases": latest_google_play_purchases,
        "sales_contact_url": SALES_CONTACT_URL if BILLING_PROVIDER != "google_play" else "",
        "play_billing_required": BILLING_PROVIDER == "google_play",
        "google_play": {
            "package_name": GOOGLE_PLAY_PACKAGE_NAME,
            "configured": google_play_configured(),
            "auto_acknowledge": GOOGLE_PLAY_AUTO_ACKNOWLEDGE,
            "products": {code: google_play_product_id_for_plan(code) for code in ("starter", "standard", "pro")},
        },
    }


def require_billing_capacity(site_code: str, metric: str, next_count: int) -> None:
    if not BILLING_ENFORCEMENT_ENABLED:
        return
    status = billing_status_for_site(site_code)
    billing = status["billing"]
    plan = status["current_plan"]
    if billing["status"] not in {"trialing", "active"}:
        raise HTTPException(status_code=402, detail="요금제 결제가 필요합니다.")
    limit = int(plan.get(f"{metric}_limit") or 0)
    if limit > 0 and next_count > limit:
        raise HTTPException(status_code=402, detail="현재 요금제 한도를 초과했습니다. 업그레이드가 필요합니다.")


def parse_google_time(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def latest_google_expiry(line_items: list[dict[str, Any]]) -> str | None:
    values = [str(item.get("expiryTime") or "") for item in line_items if item.get("expiryTime")]
    if not values:
        return None
    return max(values, key=lambda item: parse_google_time(item) or datetime.min.replace(tzinfo=timezone.utc))


def google_subscription_status(subscription: dict[str, Any]) -> tuple[str, bool]:
    state = str(subscription.get("subscriptionState") or "")
    line_items = subscription.get("lineItems") if isinstance(subscription.get("lineItems"), list) else []
    expires_at = latest_google_expiry(line_items)
    expiry_dt = parse_google_time(expires_at)
    expires_in_future = expiry_dt is None or expiry_dt > datetime.now(timezone.utc)

    if state in {"SUBSCRIPTION_STATE_ACTIVE", "SUBSCRIPTION_STATE_IN_GRACE_PERIOD"}:
        return "active", True
    if state == "SUBSCRIPTION_STATE_CANCELED" and expires_in_future:
        return "active", True
    if state in {"SUBSCRIPTION_STATE_ON_HOLD", "SUBSCRIPTION_STATE_PAUSED", "SUBSCRIPTION_STATE_PENDING"}:
        return "past_due", False
    return "cancelled", False


def google_play_authorized_session():
    if not google_play_configured():
        raise HTTPException(status_code=503, detail="Google Play 서비스 계정 설정이 필요합니다.")
    try:
        from google.auth.transport.requests import AuthorizedSession
        from google.oauth2 import service_account
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="Google Play 검증 라이브러리가 설치되어 있지 않습니다.") from exc

    scopes = ["https://www.googleapis.com/auth/androidpublisher"]
    try:
        if GOOGLE_PLAY_SERVICE_ACCOUNT_JSON:
            info = json.loads(GOOGLE_PLAY_SERVICE_ACCOUNT_JSON)
            credentials = service_account.Credentials.from_service_account_info(info, scopes=scopes)
        else:
            credentials = service_account.Credentials.from_service_account_file(GOOGLE_PLAY_SERVICE_ACCOUNT_FILE, scopes=scopes)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Google Play 서비스 계정 설정을 읽을 수 없습니다: {exc}") from exc
    return AuthorizedSession(credentials)


def fetch_google_play_subscription(package_name: str, purchase_token: str) -> dict[str, Any]:
    session = google_play_authorized_session()
    url = (
        "https://androidpublisher.googleapis.com/androidpublisher/v3/applications/"
        f"{quote(package_name, safe='')}/purchases/subscriptionsv2/tokens/{quote(purchase_token, safe='')}"
    )
    response = session.get(url, timeout=20)
    if response.status_code == 404:
        raise HTTPException(status_code=400, detail="Google Play 구매 토큰을 찾을 수 없습니다.")
    if response.status_code >= 400:
        raise HTTPException(status_code=502, detail=f"Google Play 구매 검증 실패: {response.status_code}")
    return response.json()


def acknowledge_google_play_subscription(package_name: str, product_id: str, purchase_token: str) -> bool:
    session = google_play_authorized_session()
    url = (
        "https://androidpublisher.googleapis.com/androidpublisher/v3/applications/"
        f"{quote(package_name, safe='')}/purchases/subscriptions/{quote(product_id, safe='')}"
        f"/tokens/{quote(purchase_token, safe='')}:acknowledge"
    )
    response = session.post(url, json={}, timeout=20)
    if response.status_code in {200, 204}:
        return True
    if response.status_code == 409:
        return True
    raise HTTPException(status_code=502, detail=f"Google Play 구독 승인 실패: {response.status_code}")


def save_google_play_purchase(
    *,
    con,
    site_code: str,
    username: str,
    package_name: str,
    product_id: str,
    plan: str,
    purchase_token: str,
    subscription: dict[str, Any],
    status: str,
) -> None:
    line_items = subscription.get("lineItems") if isinstance(subscription.get("lineItems"), list) else []
    expires_at = latest_google_expiry(line_items)
    order_id = subscription.get("latestOrderId")
    subscription_state = subscription.get("subscriptionState")
    acknowledgement_state = subscription.get("acknowledgementState")
    raw_response = json.dumps(subscription, ensure_ascii=False, sort_keys=True)
    con.execute(
        """
        INSERT INTO google_play_purchases
        (site_code, username, package_name, product_id, plan, purchase_token, order_id,
         subscription_state, acknowledgement_state, expires_at, raw_response_json, verified_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
        ON CONFLICT(purchase_token) DO UPDATE SET
          site_code = excluded.site_code,
          username = excluded.username,
          package_name = excluded.package_name,
          product_id = excluded.product_id,
          plan = excluded.plan,
          order_id = excluded.order_id,
          subscription_state = excluded.subscription_state,
          acknowledgement_state = excluded.acknowledgement_state,
          expires_at = excluded.expires_at,
          raw_response_json = excluded.raw_response_json,
          verified_at = datetime('now'),
          updated_at = datetime('now')
        """,
        (
            site_code,
            username,
            package_name,
            product_id,
            plan,
            purchase_token,
            order_id,
            subscription_state,
            acknowledgement_state,
            expires_at,
            raw_response,
        ),
    )
    current = con.execute("SELECT plan, payment_provider FROM site_billing WHERE site_code = ?", (site_code,)).fetchone()
    should_update_billing = status == "active" or (
        current and current["payment_provider"] == "google_play" and current["plan"] == plan
    )
    if should_update_billing:
        con.execute(
            """
            UPDATE site_billing
            SET plan = ?, status = ?, current_period_ends_at = ?, payment_provider = 'google_play',
                external_customer_id = ?, updated_at = datetime('now')
            WHERE site_code = ?
            """,
            (plan, status, expires_at, order_id, site_code),
        )


def apply_google_play_subscription_verification(
    *,
    site_code: str,
    username: str,
    package_name: str,
    product_id: str,
    purchase_token: str,
) -> dict[str, Any]:
    plan = plan_for_google_play_product(product_id)
    subscription = fetch_google_play_subscription(package_name, purchase_token)
    line_items = subscription.get("lineItems") if isinstance(subscription.get("lineItems"), list) else []
    line_product_ids = {str(item.get("productId")) for item in line_items if item.get("productId")}
    if line_product_ids and product_id not in line_product_ids:
        raise HTTPException(status_code=400, detail="구매 토큰의 상품 ID가 선택한 요금제와 다릅니다.")

    status, entitlement_active = google_subscription_status(subscription)
    acknowledgement_state = str(subscription.get("acknowledgementState") or "")
    acknowledged = acknowledgement_state == "ACKNOWLEDGEMENT_STATE_ACKNOWLEDGED"
    if GOOGLE_PLAY_AUTO_ACKNOWLEDGE and entitlement_active and not acknowledged:
        acknowledge_google_play_subscription(package_name, product_id, purchase_token)
        acknowledged = True
        acknowledgement_state = "ACKNOWLEDGEMENT_STATE_ACKNOWLEDGED"
        subscription["acknowledgementState"] = acknowledgement_state

    with connect() as con:
        ensure_site_billing_row(con, site_code)
        save_google_play_purchase(
            con=con,
            site_code=site_code,
            username=username,
            package_name=package_name,
            product_id=product_id,
            plan=plan,
            purchase_token=purchase_token,
            subscription=subscription,
            status=status,
        )
        con.commit()

    return {
        "verified": True,
        "entitlement_active": entitlement_active,
        "plan": plan,
        "google_play": {
            "package_name": package_name,
            "product_id": product_id,
            "subscription_state": subscription.get("subscriptionState"),
            "acknowledgement_state": acknowledgement_state,
            "acknowledged": acknowledged,
            "order_id": subscription.get("latestOrderId"),
            "expires_at": latest_google_expiry(line_items),
        },
        "billing_status": billing_status_for_site(site_code),
    }


def current_site_code(request: Request) -> str:
    session = read_session(request)
    if session and session.get("sc"):
        return normalize_site_code(session["sc"])
    return normalize_site_code(DEFAULT_SITE_CODE)


def ensure_ready() -> None:
    global _app_ready
    if _app_ready:
        return
    with _ready_lock:
        if _app_ready:
            return
        init_db()
        seed_users()
        maybe_seed_demo()
        auto_sync_registry()
        _app_ready = True


def lookup_vehicle(site_code: str, plate: str) -> dict[str, Any] | None:
    ensure_ready()
    normalized = normalize_plate(plate)
    with connect() as con:
        row = con.execute(
            "SELECT * FROM vehicles WHERE site_code = ? AND plate = ?",
            (site_code, normalized),
        ).fetchone()
    return dict(row) if row else None


def lookup_vehicles_by_suffix(site_code: str, suffix: str) -> list[dict[str, Any]]:
    ensure_ready()
    with connect() as con:
        rows = con.execute(
            """
            SELECT *
            FROM vehicles
            WHERE site_code = ? AND plate LIKE ?
            ORDER BY plate
            """,
            (site_code, f"%{suffix}"),
        ).fetchall()
    return [dict(row) for row in rows]


def is_suffix_plate_query(value: str) -> bool:
    raw = re.sub(r"\s+", "", str(value or ""))
    return len(raw) == 4 and raw.isdigit()


def build_check_match(plate: str, vehicle: dict[str, Any] | None) -> CheckMatch:
    verdict: PlateVerdict = evaluate_vehicle_row(vehicle)
    normalized = normalize_plate(plate)
    return CheckMatch(
        plate=normalized,
        verdict=verdict.verdict,
        message=verdict.message,
        unit=verdict.unit,
        building=(vehicle or {}).get("building"),
        unit_number=(vehicle or {}).get("unit_number"),
        owner_name=verdict.owner_name,
        phone=(vehicle or {}).get("phone"),
        status=verdict.status,
        valid_from=verdict.valid_from,
        valid_to=verdict.valid_to,
    )


def choose_best_scan_candidate(site_code: str, raw_ocr_text: str | None, manual_plate: str | None, candidates: list[str]) -> tuple[str | None, list[str]]:
    manual_normalized = normalize_plate(manual_plate)
    learned_candidates, learning_scores = get_learning_candidates(site_code, raw_ocr_text, candidates)
    source_candidates: list[str] = []
    if manual_normalized:
        source_candidates.append(manual_normalized)
    source_candidates.extend(learned_candidates)
    source_candidates.extend(candidates)

    normalized_candidates: list[str] = []
    seen: set[str] = set()

    for candidate in source_candidates:
        normalized = normalize_plate(candidate)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        normalized_candidates.append(normalized)

    if not normalized_candidates:
        return None, []

    ranked: list[tuple[str, float]] = []
    for index, candidate in enumerate(normalized_candidates):
        vehicle = lookup_vehicle(site_code, candidate)
        verdict = evaluate_vehicle_row(vehicle)
        score = learning_scores.get(candidate, 0.0)
        if candidate == manual_normalized:
            score += 1000.0
        if verdict.verdict != "UNREGISTERED":
            score += 220.0
            if verdict.verdict == "OK":
                score += 40.0
        score += max(0.0, 30.0 - (index * 2.0))
        ranked.append((candidate, score))

    ranked.sort(key=lambda item: item[1], reverse=True)
    ordered = [candidate for candidate, _ in ranked[:8]]
    return ordered[0], ordered


def build_check_response(site_code: str, plate: str) -> CheckResponse:
    requested = str(plate or "").strip()
    normalized = normalize_plate(requested)
    if not normalized and not is_suffix_plate_query(requested):
        raise HTTPException(status_code=400, detail="차량번호를 입력해 주세요.")

    if is_suffix_plate_query(requested):
        suffix_matches = [
            build_check_match(row["plate"], row)
            for row in lookup_vehicles_by_suffix(site_code, requested)
        ]
        suffix_matches.sort(
            key=lambda item: (
                {"OK": 0, "TEMP": 1, "EXPIRED": 2, "BLOCKED": 3, "UNREGISTERED": 4}.get(item.verdict, 9),
                item.plate,
            )
        )
        if not suffix_matches:
            unmatched = build_check_match(requested, None)
            return CheckResponse(
                site_code=site_code,
                requested_plate=requested,
                match_mode="suffix",
                match_count=0,
                match_index=0,
                matches=[],
                **unmatched.model_dump(),
            )

        primary = suffix_matches[0]
        return CheckResponse(
            site_code=site_code,
            requested_plate=requested,
            match_mode="suffix",
            match_count=len(suffix_matches),
            match_index=0,
            matches=suffix_matches,
            **primary.model_dump(),
        )

    match = build_check_match(normalized, lookup_vehicle(site_code, normalized))
    return CheckResponse(
        site_code=site_code,
        requested_plate=requested or normalized,
        match_mode="exact",
        match_count=1,
        match_index=0,
        matches=[],
        **match.model_dump(),
    )


def normalize_username(value: str | None) -> str:
    username = str(value or "").strip().lower()
    if not USERNAME_PATTERN.fullmatch(username):
        raise HTTPException(status_code=400, detail="아이디는 영문 소문자, 숫자, 마침표(.), 밑줄(_), 하이픈(-)만 사용해 3~32자로 입력해 주세요.")
    return username


def normalize_user_role(value: str | None) -> str:
    role = str(value or "").strip().lower()
    if role not in VALID_USER_ROLES:
        labels = ", ".join(ROLE_LABELS.values())
        raise HTTPException(status_code=400, detail=f"권한은 {labels} 중 하나여야 합니다.")
    return role


def normalize_cctv_status(value: str | None) -> str:
    status = str(value or "").strip().lower()
    if status not in CCTV_STATUSES:
        raise HTTPException(status_code=400, detail="CCTV 요청 상태를 다시 확인해 주세요.")
    return status


def require_form_text(value: str | None, label: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail=f"{label}을 입력해 주세요.")
    return text


def validate_cctv_time_range(search_start_time: str, search_end_time: str) -> None:
    if search_end_time < search_start_time:
        raise HTTPException(status_code=400, detail="검색 끝 시간은 시작 시간 이후로 입력해 주세요.")


def normalize_history_datetime(value: str | None, *, end_of_range: bool = False) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    text = text.replace("T", " ")
    if len(text) == 10:
        return f"{text} {'23:59:59' if end_of_range else '00:00:00'}"
    if len(text) == 16:
        return f"{text}:{'59' if end_of_range else '00'}"
    return text


def build_enforcement_history_query(
    site_code: str,
    *,
    q: str = "",
    verdict: str = "",
    date_from: str = "",
    date_to: str = "",
) -> tuple[str, list[Any]]:
    where = ["e.site_code = ?"]
    params: list[Any] = [site_code]

    query = str(q or "").strip()
    if query:
        normalized_plate = normalize_plate(query)
        like = f"%{query}%"
        plate_like = f"%{normalized_plate or query}%"
        where.append(
            """
            (
              e.plate LIKE ?
              OR COALESCE(e.unit, '') LIKE ?
              OR COALESCE(e.owner_name, '') LIKE ?
              OR COALESCE(e.inspector, '') LIKE ?
              OR COALESCE(e.location, '') LIKE ?
              OR COALESCE(e.memo, '') LIKE ?
            )
            """
        )
        params.extend([plate_like, like, like, like, like, like])

    normalized_verdict = str(verdict or "").strip().upper()
    if normalized_verdict:
        where.append("e.verdict = ?")
        params.append(normalized_verdict)

    normalized_from = normalize_history_datetime(date_from)
    if normalized_from:
        where.append("e.created_at >= ?")
        params.append(normalized_from)

    normalized_to = normalize_history_datetime(date_to, end_of_range=True)
    if normalized_to:
        where.append("e.created_at <= ?")
        params.append(normalized_to)

    return " AND ".join(where), params


def fetch_enforcement_history_rows(
    site_code: str,
    *,
    q: str = "",
    verdict: str = "",
    date_from: str = "",
    date_to: str = "",
    limit: int = 1000,
    offset: int = 0,
) -> list[dict[str, Any]]:
    where_sql, params = build_enforcement_history_query(
        site_code,
        q=q,
        verdict=verdict,
        date_from=date_from,
        date_to=date_to,
    )
    params.extend([limit, offset])
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT
              e.id,
              e.plate,
              e.verdict,
              e.verdict_message,
              e.unit,
              e.owner_name,
              e.vehicle_status,
              e.inspector,
              e.location,
              e.memo,
              e.photo_path,
              e.lat,
              e.lng,
              e.created_at,
              v.phone,
              v.building,
              v.unit_number
            FROM enforcement_events e
            LEFT JOIN vehicles v ON v.site_code = e.site_code AND v.plate = e.plate
            WHERE {where_sql}
            ORDER BY e.id DESC
            LIMIT ? OFFSET ?
            """,
            params,
        ).fetchall()
    return [dict(row) for row in rows]


def fetch_enforcement_event(site_code: str, event_id: int) -> dict[str, Any] | None:
    with connect() as con:
        row = con.execute(
            """
            SELECT
              e.id,
              e.plate,
              e.verdict,
              e.verdict_message,
              e.unit,
              e.owner_name,
              e.vehicle_status,
              e.inspector,
              e.location,
              e.memo,
              e.photo_path,
              e.lat,
              e.lng,
              e.created_at,
              v.phone,
              v.building,
              v.unit_number
            FROM enforcement_events e
            LEFT JOIN vehicles v ON v.site_code = e.site_code AND v.plate = e.plate
            WHERE e.site_code = ? AND e.id = ?
            """,
            (site_code, event_id),
        ).fetchone()
    return dict(row) if row else None


def require_enforcement_event(site_code: str, event_id: int) -> dict[str, Any]:
    row = fetch_enforcement_event(site_code, event_id)
    if not row:
        raise HTTPException(status_code=404, detail="단속 기록을 찾을 수 없습니다.")
    return row


def normalize_new_password(value: str | None, *, required: bool) -> str | None:
    if value is None:
        if required:
            raise HTTPException(status_code=400, detail="비밀번호를 입력해 주세요.")
        return None

    password = str(value)
    if not password:
        if required:
            raise HTTPException(status_code=400, detail="비밀번호를 입력해 주세요.")
        return None
    if password != password.strip():
        raise HTTPException(status_code=400, detail="비밀번호 앞뒤 공백은 사용할 수 없습니다.")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="비밀번호는 8자 이상이어야 합니다.")
    return password


def user_public_dict(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    return {
        "site_code": row["site_code"],
        "username": row["username"],
        "role": row["role"],
        "role_label": ROLE_LABELS.get(row["role"], row["role"]),
        "can_manage_vehicles": bool(row.get("can_manage_vehicles", 0)),
        "created_at": row["created_at"],
    }


def require_existing_user(con, site_code: str, username: str) -> dict[str, Any]:
    row = con.execute(
        "SELECT site_code, username, role, COALESCE(can_manage_vehicles, 0) AS can_manage_vehicles, created_at FROM users WHERE site_code = ? AND username = ?",
        (normalize_site_code(site_code), username),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="해당 사용자를 찾을 수 없습니다.")
    return dict(row)


def can_manage_vehicle_registry(request: Request) -> bool:
    session = require_role(request, VIEW_ROLES)
    if session.get("r") == "admin":
        return True
    site_code = current_site_code(request)
    with connect() as con:
        row = con.execute(
            "SELECT COALESCE(can_manage_vehicles, 0) AS can_manage_vehicles FROM users WHERE site_code = ? AND username = ?",
            (site_code, session.get("u")),
        ).fetchone()
    return bool(row and row["can_manage_vehicles"])


def require_vehicle_manager(request: Request) -> dict[str, Any]:
    session = require_role(request, VIEW_ROLES)
    if session.get("r") == "admin":
        return session
    site_code = current_site_code(request)
    with connect() as con:
        row = con.execute(
            "SELECT COALESCE(can_manage_vehicles, 0) AS can_manage_vehicles FROM users WHERE site_code = ? AND username = ?",
            (site_code, session.get("u")),
        ).fetchone()
    if not row or not row["can_manage_vehicles"]:
        raise HTTPException(status_code=403, detail="등록차량 DB를 관리할 권한이 없습니다.")
    return session


def vehicle_row_dict(row: dict[str, Any] | Any | None) -> dict[str, Any] | None:
    if not row:
        return None
    data = dict(row)
    data["manual_override"] = bool(data.get("manual_override", 0))
    return data


def vehicle_payload_values(payload: VehicleUpsertRequest) -> dict[str, Any]:
    plate = normalize_plate(payload.plate)
    if not plate:
        raise HTTPException(status_code=400, detail="차량번호를 입력해 주세요.")
    status = str(payload.status or "active").strip().lower() or "active"
    return {
        "plate": plate,
        "unit": str(payload.unit or "").strip() or None,
        "building": str(payload.building or "").strip() or None,
        "unit_number": str(payload.unit_number or "").strip() or None,
        "owner_name": str(payload.owner_name or "").strip() or None,
        "phone": str(payload.phone or "").strip() or None,
        "status": status,
        "valid_from": str(payload.valid_from or "").strip() or None,
        "valid_to": str(payload.valid_to or "").strip() or None,
        "note": str(payload.note or "").strip() or None,
    }


def log_vehicle_change(con, site_code: str, username: str | None, action: str, plate: str | None, before: dict[str, Any] | None, after: dict[str, Any] | None) -> None:
    con.execute(
        """
        INSERT INTO vehicle_change_logs(site_code, username, action, plate, before_json, after_json)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            site_code,
            username,
            action,
            plate,
            json.dumps(before, ensure_ascii=False, default=str) if before else None,
            json.dumps(after, ensure_ascii=False, default=str) if after else None,
        ),
    )


def create_vehicle_backup(con, site_code: str, username: str | None, backup_name: str | None = None) -> dict[str, Any]:
    rows = [
        dict(row)
        for row in con.execute(
            """
            SELECT site_code, plate, unit, building, unit_number, owner_name, phone, status, valid_from, valid_to, note, source_file, source_sheet, manual_override, deleted_at, updated_at
            FROM vehicles
            WHERE site_code = ?
            ORDER BY plate
            """,
            (site_code,),
        ).fetchall()
    ]
    name = backup_name or f"{site_code}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    cur = con.execute(
        """
        INSERT INTO vehicle_backups(site_code, backup_name, vehicles_json, vehicles_count, created_by)
        VALUES (?, ?, ?, ?, ?)
        """,
        (site_code, name, json.dumps(rows, ensure_ascii=False, default=str), len(rows), username),
    )
    return {"id": cur.lastrowid, "backup_name": name, "vehicles_count": len(rows)}


def ensure_not_last_admin(con, site_code: str, username: str, *, next_role: str | None = None, deleting: bool = False) -> None:
    normalized_site = normalize_site_code(site_code)
    target = require_existing_user(con, normalized_site, username)
    if target["role"] != "admin":
        return

    if not deleting and (next_role is None or next_role == "admin"):
        return

    admin_count = con.execute(
        "SELECT COUNT(*) AS cnt FROM users WHERE site_code = ? AND role = 'admin'",
        (normalized_site,),
    ).fetchone()["cnt"]
    if int(admin_count) <= 1:
        raise HTTPException(status_code=400, detail="마지막 관리자 계정은 삭제하거나 다른 권한으로 변경할 수 없습니다.")


def save_photo(photo: UploadFile, site_code: str) -> str | None:
    if not photo.filename:
        return None
    target_dir = site_upload_dir(site_code)
    target_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(photo.filename).suffix.lower() or ".jpg"
    name = f"{uuid.uuid4().hex}{suffix}"
    file_path = target_dir / name
    with file_path.open("wb") as target:
        shutil.copyfileobj(photo.file, target)
    return site_upload_url(site_code, name)


def save_photo_bytes(filename: str | None, payload: bytes, site_code: str) -> str:
    if not filename:
        raise HTTPException(status_code=400, detail="사진 파일을 선택해 주세요.")
    if not payload:
        raise HTTPException(status_code=400, detail="사진 파일이 비어 있습니다.")
    target_dir = site_upload_dir(site_code)
    target_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(filename).suffix.lower() or ".jpg"
    name = f"{uuid.uuid4().hex}{suffix}"
    (target_dir / name).write_bytes(payload)
    return site_upload_url(site_code, name)


def save_site_setting_image(filename: str | None, payload: bytes, site_code: str) -> str:
    if not filename:
        raise HTTPException(status_code=400, detail="이미지 파일을 선택해 주세요.")
    if not payload:
        raise HTTPException(status_code=400, detail="이미지 파일이 비어 있습니다.")
    suffix = Path(filename).suffix.lower() or ".jpg"
    if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        raise HTTPException(status_code=400, detail="이미지는 jpg, png, webp, gif 형식만 사용할 수 있습니다.")
    target_dir = site_upload_dir(site_code) / "settings"
    target_dir.mkdir(parents=True, exist_ok=True)
    name = f"capture-placeholder-{uuid.uuid4().hex}{suffix}"
    (target_dir / name).write_bytes(payload)
    return site_upload_url(site_code, f"settings/{name}")


def site_settings_dict(site_code: str) -> dict[str, Any]:
    normalized_site = normalize_site_code(site_code)
    with connect() as con:
        row = con.execute("SELECT * FROM site_settings WHERE site_code = ?", (normalized_site,)).fetchone()
    return {
        "site_code": normalized_site,
        "capture_placeholder_image_url": row["capture_placeholder_image_url"] if row else None,
        "updated_at": row["updated_at"] if row else None,
    }


def cctv_request_dict(row: dict[str, Any] | Any) -> dict[str, Any]:
    data = dict(row)
    start_time = data.get("search_start_time") or data.get("search_time")
    end_time = data.get("search_end_time") or start_time
    data["search_start_time"] = start_time
    data["search_end_time"] = end_time
    data["status_label"] = CCTV_STATUS_LABELS.get(data.get("status"), data.get("status") or "-")
    return data


def require_cctv_request(con, site_code: str, request_id: int) -> dict[str, Any]:
    row = con.execute(
        "SELECT * FROM cctv_search_requests WHERE site_code = ? AND id = ?",
        (site_code, request_id),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="CCTV 검색요청을 찾을 수 없습니다.")
    return dict(row)


def can_edit_cctv_request(session: dict[str, Any], row: dict[str, Any]) -> bool:
    username = session.get("u")
    role = session.get("r")
    return role in CCTV_ASSIGNMENT_ROLES or row.get("requester_username") == username or row.get("assigned_to") == username


def can_delete_cctv_request(session: dict[str, Any], row: dict[str, Any]) -> bool:
    return session.get("r") in CCTV_ASSIGNMENT_ROLES or row.get("requester_username") == session.get("u")


def normalize_cctv_assignee(con, site_code: str, username: str | None) -> str | None:
    assignee = str(username or "").strip()
    if not assignee:
        return None
    row = con.execute(
        "SELECT username, role FROM users WHERE site_code = ? AND username = ?",
        (normalize_site_code(site_code), assignee),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=400, detail="담당자 계정을 찾을 수 없습니다.")
    if row["role"] not in VALID_USER_ROLES:
        raise HTTPException(status_code=400, detail="담당자 권한을 다시 확인해 주세요.")
    return row["username"]


def auto_sync_registry() -> None:
    with connect() as con:
        site_codes = [row["site_code"] for row in con.execute("SELECT site_code FROM sites ORDER BY site_code").fetchall()]

    for site_code in site_codes:
        source_dir = site_import_dir(site_code)
        excel_files = [
            path
            for path in source_dir.iterdir()
            if source_dir.exists() and path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"} and not path.name.startswith("~$")
        ] if source_dir.exists() else []
        if not excel_files:
            continue
        try:
            sync_registry_from_dir(source_dir, site_code)
        except Exception as exc:
            print(f"[startup] registry sync failed for {site_code}: {exc}")


@app.on_event("startup")
def on_startup() -> None:
    ensure_ready()


@app.get("/health")
def health() -> dict[str, bool]:
    return {"ok": True}


@app.get("/")
def root(request: Request):
    if read_session(request):
        return RedirectResponse(url=app_url("/field"), status_code=302)
    return RedirectResponse(url=app_url("/login"), status_code=302)


@app.head("/", include_in_schema=False)
def root_head():
    return Response(status_code=302, headers={"Location": app_url("/login")})


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return Response(status_code=204)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if read_session(request):
        return RedirectResponse(url=app_url("/field"), status_code=302)
    if not LOCAL_LOGIN_ENABLED:
        return HTMLResponse("<h2>통합 로그인 모드입니다.</h2>", status_code=403)
    return render_login_page(request, site_code=DEFAULT_SITE_CODE)


@app.post("/login")
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    site_code: str | None = Form(None),
):
    if not LOCAL_LOGIN_ENABLED:
        raise HTTPException(status_code=403, detail="통합 로그인 전용입니다.")

    ensure_ready()
    raw_site_code = str(site_code or DEFAULT_SITE_CODE).strip()
    try:
        user_site_code = normalize_login_site_code(raw_site_code)
    except HTTPException:
        return render_login_page(
            request,
            status_code=400,
            username=str(username or "").strip(),
            site_code=raw_site_code,
            error=LOGIN_SITE_FORMAT_MESSAGE,
        )
    try:
        user_name = normalize_username(username)
    except HTTPException:
        return render_login_page(
            request,
            status_code=400,
            username=str(username or "").strip(),
            site_code=user_site_code,
            error=LOGIN_FORMAT_MESSAGE,
        )
    with connect() as con:
        row = con.execute(
            "SELECT * FROM users WHERE site_code = ? AND username = ?",
            (user_site_code, user_name),
        ).fetchone()
    if not row or not pbkdf2_verify(password, row["pw_hash"]):
        return render_login_page(
            request,
            status_code=401,
            username=user_name,
            site_code=user_site_code,
            error=LOGIN_INVALID_MESSAGE,
        )

    token = make_session(user_name, row["role"], user_site_code)
    response = RedirectResponse(url=app_url("/field"), status_code=302)
    response.set_cookie(
        COOKIE_NAME,
        token,
        httponly=True,
        samesite="lax",
        path=session_cookie_path(),
    )
    return response


@app.post("/logout")
def logout():
    response = RedirectResponse(url=app_url("/login"), status_code=302)
    response.delete_cookie(COOKIE_NAME, path=session_cookie_path())
    return response


@app.get("/field", response_class=HTMLResponse)
def field_page(request: Request):
    ensure_ready()
    session = require_role(request, VIEW_ROLES)
    site_code = normalize_site_code(session.get("sc"))
    can_manage_vehicles = can_manage_vehicle_registry(request)
    return templates.TemplateResponse(
        request=request,
        name="field.html",
        context={
            "app_title": APP_TITLE,
            "site_code": site_code,
            "site_name": site_name_for_code(site_code),
            "role": session.get("r"),
            "username": session.get("u"),
            "is_admin": session.get("r") == "admin",
            "can_manage_vehicles": can_manage_vehicles,
            "import_dir": str(site_import_dir(site_code)),
            "ocr_provider": os.getenv("PARKING_OCR_PROVIDER", "tesseract"),
        },
    )


@app.get("/api/me")
def api_me(request: Request):
    ensure_ready()
    session = require_role(request, VIEW_ROLES)
    return {
        "username": session.get("u"),
        "role": session.get("r"),
        "site_code": normalize_site_code(session.get("sc")),
        "site_name": site_name_for_code(session.get("sc")),
        "can_manage_vehicles": can_manage_vehicle_registry(request),
    }


@app.get("/api/billing/status")
def api_billing_status(request: Request):
    ensure_ready()
    require_role(request, {"admin"})
    return billing_status_for_site(current_site_code(request))


@app.post("/api/billing/inquiries")
def api_billing_inquiry_create(request: Request, payload: BillingInquiryRequest):
    ensure_ready()
    require_role(request, {"admin"})
    site_code = current_site_code(request)
    requested_plan = normalize_billing_plan(payload.requested_plan)
    contact_name = normalize_billing_text(payload.contact_name, "담당자명", 40)
    contact_phone = normalize_billing_text(payload.contact_phone, "연락처", 40)
    contact_email = normalize_billing_text(payload.contact_email, "이메일", 120)
    message = normalize_billing_text(payload.message, "문의 내용", 500)

    if not any([contact_name, contact_phone, contact_email, message]):
        raise HTTPException(status_code=400, detail="연락처 또는 문의 내용을 하나 이상 입력해 주세요.")

    with connect() as con:
        ensure_site_billing_row(con, site_code)
        con.execute(
            """
            INSERT INTO billing_inquiries(site_code, requested_plan, contact_name, contact_phone, contact_email, message)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (site_code, requested_plan, contact_name, contact_phone, contact_email, message),
        )
        con.commit()
    return billing_status_for_site(site_code)


@app.get("/api/billing/google-play/config")
def api_google_play_billing_config(request: Request):
    ensure_ready()
    require_role(request, {"admin"})
    return {
        "enabled": BILLING_PROVIDER == "google_play",
        "configured": google_play_configured(),
        "package_name": GOOGLE_PLAY_PACKAGE_NAME,
        "products": {code: google_play_product_id_for_plan(code) for code in ("starter", "standard", "pro")},
    }


@app.post("/api/billing/google-play/verify")
def api_google_play_billing_verify(request: Request, payload: GooglePlayVerifyRequest):
    ensure_ready()
    session = require_role(request, {"admin"})
    site_code = current_site_code(request)
    username = str(session.get("u") or "")
    product_id = normalize_google_play_product_id(payload.product_id)
    purchase_token = normalize_purchase_token(payload.purchase_token)
    package_name = str(payload.package_name or GOOGLE_PLAY_PACKAGE_NAME).strip() or GOOGLE_PLAY_PACKAGE_NAME
    if package_name != GOOGLE_PLAY_PACKAGE_NAME:
        raise HTTPException(status_code=400, detail="Google Play 패키지명이 서버 설정과 다릅니다.")

    return apply_google_play_subscription_verification(
        site_code=site_code,
        username=username,
        package_name=package_name,
        product_id=product_id,
        purchase_token=purchase_token,
    )


@app.post("/api/billing/google-play/rtdn")
async def api_google_play_billing_rtdn(request: Request):
    ensure_ready()
    if not GOOGLE_PLAY_RTDN_TOKEN:
        raise HTTPException(status_code=404, detail="Google Play RTDN 수신이 설정되지 않았습니다.")
    provided_token = request.headers.get("x-parking-webhook-token") or request.query_params.get("token")
    if provided_token != GOOGLE_PLAY_RTDN_TOKEN:
        raise HTTPException(status_code=403, detail="Google Play RTDN 토큰이 올바르지 않습니다.")

    try:
        body = await request.json()
        encoded_data = body.get("message", {}).get("data", "")
        notification = json.loads(base64.b64decode(encoded_data).decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Google Play RTDN 메시지를 해석할 수 없습니다.") from exc

    subscription = notification.get("subscriptionNotification") or {}
    product_id = normalize_google_play_product_id(subscription.get("subscriptionId"))
    purchase_token = normalize_purchase_token(subscription.get("purchaseToken"))

    with connect() as con:
        row = con.execute(
            """
            SELECT site_code, username, package_name
            FROM google_play_purchases
            WHERE purchase_token = ?
            ORDER BY verified_at DESC
            LIMIT 1
            """,
            (purchase_token,),
        ).fetchone()
    if not row:
        return {"ignored": True, "reason": "unknown_purchase_token"}

    return apply_google_play_subscription_verification(
        site_code=row["site_code"],
        username=row["username"],
        package_name=row["package_name"] or GOOGLE_PLAY_PACKAGE_NAME,
        product_id=product_id,
        purchase_token=purchase_token,
    )


@app.get("/api/users")
def api_users_list(request: Request, q: str = "", role: str = "", limit: int = 50, offset: int = 0):
    ensure_ready()
    require_role(request, {"admin"})
    site_code = current_site_code(request)
    limit = min(max(limit, 1), 100)
    offset = max(offset, 0)

    where = ["site_code = ?"]
    params: list[Any] = [site_code]
    query = str(q or "").strip().lower()
    if query:
        where.append("username LIKE ?")
        params.append(f"%{query}%")

    normalized_role = str(role or "").strip().lower()
    if normalized_role:
        normalized_role = normalize_user_role(normalized_role)
        where.append("role = ?")
        params.append(normalized_role)

    params.extend([limit, offset])
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT site_code, username, role, COALESCE(can_manage_vehicles, 0) AS can_manage_vehicles, created_at
            FROM users
            WHERE {' AND '.join(where)}
            ORDER BY {role_order_case()}, username
            LIMIT ? OFFSET ?
            """,
            params,
        ).fetchall()
    return [user_public_dict(dict(row)) for row in rows]


@app.post("/api/users")
def api_users_create(request: Request, payload: UserCreateRequest):
    ensure_ready()
    require_role(request, {"admin"})
    site_code = current_site_code(request)

    username = normalize_username(payload.username)
    role = normalize_user_role(payload.role)
    password = normalize_new_password(payload.password, required=True)

    if BILLING_ENFORCEMENT_ENABLED:
        current_usage = billing_status_for_site(site_code)["usage"]["users"]
        require_billing_capacity(site_code, "users", int(current_usage) + 1)

    with connect() as con:
        exists = con.execute(
            "SELECT 1 FROM users WHERE site_code = ? AND username = ?",
            (site_code, username),
        ).fetchone()
        if exists:
            raise HTTPException(status_code=409, detail="이미 사용 중인 아이디입니다.")

        con.execute(
            "INSERT INTO users(site_code, username, pw_hash, role, can_manage_vehicles) VALUES (?, ?, ?, ?, ?)",
            (site_code, username, pbkdf2_hash(password), role, 1 if payload.can_manage_vehicles else 0),
        )
        row = require_existing_user(con, site_code, username)
        con.commit()
    return user_public_dict(row)


@app.patch("/api/users/{username}")
def api_users_update(request: Request, username: str, payload: UserUpdateRequest):
    ensure_ready()
    session = require_role(request, {"admin"})
    site_code = current_site_code(request)
    normalized_username = normalize_username(username)
    next_role = normalize_user_role(payload.role) if payload.role is not None else None
    next_password = normalize_new_password(payload.password, required=False)

    with connect() as con:
        current = require_existing_user(con, site_code, normalized_username)

        if normalized_username == session.get("u") and next_role is not None and next_role != current["role"]:
            raise HTTPException(status_code=400, detail="현재 로그인한 본인 계정의 권한은 여기서 변경할 수 없습니다.")

        ensure_not_last_admin(con, site_code, normalized_username, next_role=next_role, deleting=False)

        fields: list[str] = []
        values: list[Any] = []

        if next_role is not None and next_role != current["role"]:
            fields.append("role = ?")
            values.append(next_role)

        if next_password is not None:
            fields.append("pw_hash = ?")
            values.append(pbkdf2_hash(next_password))

        if payload.can_manage_vehicles is not None:
            fields.append("can_manage_vehicles = ?")
            values.append(1 if payload.can_manage_vehicles else 0)

        if not fields:
            return user_public_dict(current)

        values.extend([site_code, normalized_username])
        con.execute(f"UPDATE users SET {', '.join(fields)} WHERE site_code = ? AND username = ?", values)
        row = require_existing_user(con, site_code, normalized_username)
        con.commit()
    return user_public_dict(row)


@app.delete("/api/users/{username}")
def api_users_delete(request: Request, username: str):
    ensure_ready()
    session = require_role(request, {"admin"})
    site_code = current_site_code(request)
    normalized_username = normalize_username(username)

    if normalized_username == session.get("u"):
        raise HTTPException(status_code=400, detail="현재 로그인한 계정은 삭제할 수 없습니다.")

    with connect() as con:
        require_existing_user(con, site_code, normalized_username)
        ensure_not_last_admin(con, site_code, normalized_username, deleting=True)
        con.execute("DELETE FROM users WHERE site_code = ? AND username = ?", (site_code, normalized_username))
        con.commit()

    return {"deleted": True, "username": normalized_username}


@app.get("/api/sites")
def api_sites_list(request: Request, q: str = "", limit: int = 50, offset: int = 0):
    ensure_ready()
    require_role(request, {"admin"})
    limit = min(max(limit, 1), 100)
    offset = max(offset, 0)
    query = str(q or "").strip()
    where = ""
    params: list[Any] = []
    if query:
        where = "WHERE s.site_code LIKE ? OR s.name LIKE ?"
        like = f"%{query}%"
        params.extend([like, like])
    params.extend([limit, offset])
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT
              s.site_code,
              s.name,
              s.created_at,
              COUNT(DISTINCT u.id) AS users_count,
              COUNT(DISTINCT v.plate) AS vehicles_count
            FROM sites s
            LEFT JOIN users u ON u.site_code = s.site_code
            LEFT JOIN vehicles v ON v.site_code = s.site_code
            {where}
            GROUP BY s.site_code, s.name, s.created_at
            ORDER BY s.site_code
            LIMIT ? OFFSET ?
            """,
            params,
        ).fetchall()
    return [site_public_dict(dict(row)) for row in rows]


@app.post("/api/sites")
def api_sites_create(request: Request, payload: SiteCreateRequest):
    ensure_ready()
    require_role(request, {"admin"})

    site_code = normalize_required_site_code(payload.site_code)
    site_name = normalize_site_name(payload.name)
    admin_username = normalize_username(payload.admin_username)
    admin_password = normalize_new_password(payload.admin_password, required=True)

    with connect() as con:
        exists = con.execute("SELECT 1 FROM sites WHERE site_code = ?", (site_code,)).fetchone()
        if exists:
            raise HTTPException(status_code=409, detail="이미 등록된 아파트 코드입니다.")

        con.execute("INSERT INTO sites(site_code, name) VALUES (?, ?)", (site_code, site_name))
        con.execute(
            "INSERT INTO users(site_code, username, pw_hash, role) VALUES (?, ?, ?, 'admin')",
            (site_code, admin_username, pbkdf2_hash(admin_password)),
        )
        con.execute(
            """
            INSERT OR IGNORE INTO site_billing(site_code, plan, status, trial_ends_at, payment_provider)
            VALUES (?, 'trial', 'trialing', date('now', '+14 days'), ?)
            """,
            (site_code, BILLING_PROVIDER),
        )
        row = con.execute(
            """
            SELECT
              s.site_code,
              s.name,
              s.created_at,
              COUNT(DISTINCT u.id) AS users_count,
              COUNT(DISTINCT v.plate) AS vehicles_count
            FROM sites s
            LEFT JOIN users u ON u.site_code = s.site_code
            LEFT JOIN vehicles v ON v.site_code = s.site_code
            WHERE s.site_code = ?
            GROUP BY s.site_code, s.name, s.created_at
            """,
            (site_code,),
        ).fetchone()
        con.commit()
    return site_public_dict(dict(row))


@app.get("/api/cctv/assignees")
def api_cctv_assignees(request: Request):
    ensure_ready()
    require_role(request, CCTV_ASSIGNMENT_ROLES)
    site_code = current_site_code(request)
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT site_code, username, role, created_at
            FROM users
            WHERE site_code = ?
            ORDER BY {role_order_case()}, username
            """,
            (site_code,),
        ).fetchall()
    return [user_public_dict(dict(row)) for row in rows]


@app.get("/api/cctv/requests")
def api_cctv_requests(request: Request, limit: int = 50, offset: int = 0):
    ensure_ready()
    session = require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    limit = min(max(limit, 1), 100)
    offset = max(offset, 0)

    base_query = """
        SELECT *
        FROM cctv_search_requests
        WHERE site_code = ?
          AND status NOT IN ('done', 'cancelled')
    """
    params: list[Any] = [site_code]
    if session.get("r") not in CCTV_ASSIGNMENT_ROLES:
        base_query += " AND (requester_username = ? OR assigned_to = ?)"
        params.extend([session.get("u"), session.get("u")])

    base_query += """
        ORDER BY
          CASE status
            WHEN 'requested' THEN 0
            WHEN 'assigned' THEN 1
            WHEN 'in_progress' THEN 2
            WHEN 'done' THEN 3
            ELSE 4
          END,
          work_weight DESC,
          search_start_time ASC,
          created_at DESC
        LIMIT ? OFFSET ?
    """
    params.extend([limit, offset])
    with connect() as con:
        rows = con.execute(base_query, params).fetchall()
    return [cctv_request_dict(row) for row in rows]


@app.post("/api/cctv/requests")
async def api_cctv_request_create(
    request: Request,
    photo: UploadFile = File(...),
    location: str = Form(...),
    search_start_time: str | None = Form(None),
    search_end_time: str | None = Form(None),
    search_time: str | None = Form(None),
    content: str = Form(...),
):
    ensure_ready()
    session = require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    normalized_location = require_form_text(location, "위치")
    normalized_search_start_time = require_form_text(search_start_time or search_time, "검색 시작 시간")
    normalized_search_end_time = require_form_text(search_end_time or search_time, "검색 끝 시간")
    validate_cctv_time_range(normalized_search_start_time, normalized_search_end_time)
    normalized_content = require_form_text(content, "요청 내용")
    if BILLING_ENFORCEMENT_ENABLED:
        current_usage = billing_status_for_site(site_code)["usage"]["monthly_cctv"]
        require_billing_capacity(site_code, "monthly_cctv", int(current_usage) + 1)
    payload = await photo.read()
    photo_path = save_photo_bytes(photo.filename, payload, site_code)

    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO cctv_search_requests
            (site_code, requester_username, photo_path, location, search_start_time, search_end_time, content)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                site_code,
                session.get("u"),
                photo_path,
                normalized_location,
                normalized_search_start_time,
                normalized_search_end_time,
                normalized_content,
            ),
        )
        row = con.execute("SELECT * FROM cctv_search_requests WHERE id = ?", (cur.lastrowid,)).fetchone()
        con.commit()
    return cctv_request_dict(row)


@app.patch("/api/cctv/requests/{request_id}")
def api_cctv_request_update(request: Request, request_id: int, payload: CctvAssignmentRequest):
    ensure_ready()
    session = require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    with connect() as con:
        current = require_cctv_request(con, site_code, request_id)
        if not can_edit_cctv_request(session, current):
            raise HTTPException(status_code=403, detail="이 CCTV 요청을 수정할 권한이 없습니다.")
        fields_set = payload.model_fields_set
        updates: list[str] = []
        values: list[Any] = []

        if "location" in fields_set:
            updates.append("location = ?")
            values.append(require_form_text(payload.location, "위치"))

        next_start_time = current.get("search_start_time") or current.get("search_time")
        next_end_time = current.get("search_end_time") or next_start_time
        if "search_start_time" in fields_set:
            next_start_time = require_form_text(payload.search_start_time, "검색 시작 시간")
        if "search_end_time" in fields_set:
            next_end_time = require_form_text(payload.search_end_time, "검색 끝 시간")
        if "search_start_time" in fields_set or "search_end_time" in fields_set:
            validate_cctv_time_range(str(next_start_time), str(next_end_time))
            updates.append("search_start_time = ?")
            values.append(next_start_time)
            updates.append("search_end_time = ?")
            values.append(next_end_time)

        if "content" in fields_set:
            updates.append("content = ?")
            values.append(require_form_text(payload.content, "요청 내용"))

        assignment_fields = {"assigned_to", "work_weight", "instruction", "status"}
        if fields_set & assignment_fields and session.get("r") not in CCTV_ASSIGNMENT_ROLES:
            raise HTTPException(status_code=403, detail="CCTV 배정 정보는 관리자만 수정할 수 있습니다.")

        if "assigned_to" in fields_set:
            assignee = normalize_cctv_assignee(con, site_code, payload.assigned_to)
            updates.append("assigned_to = ?")
            values.append(assignee)
            updates.append("assigned_by = ?")
            values.append(session.get("u") if assignee else None)
            updates.append("assigned_at = datetime('now')" if assignee else "assigned_at = NULL")
            if assignee and "status" not in fields_set and current["status"] == "requested":
                updates.append("status = ?")
                values.append("assigned")

        if "work_weight" in fields_set and payload.work_weight is not None:
            updates.append("work_weight = ?")
            values.append(payload.work_weight)

        if "instruction" in fields_set:
            updates.append("instruction = ?")
            values.append(str(payload.instruction or "").strip() or None)

        next_status: str | None = None
        if "status" in fields_set and payload.status is not None:
            next_status = normalize_cctv_status(payload.status)
            updates.append("status = ?")
            values.append(next_status)
            updates.append("completed_at = datetime('now')" if next_status == "done" else "completed_at = NULL")

        if not updates:
            return cctv_request_dict(current)

        updates.append("updated_at = datetime('now')")
        values.extend([request_id, site_code])
        con.execute(
            f"UPDATE cctv_search_requests SET {', '.join(updates)} WHERE id = ? AND site_code = ?",
            values,
        )
        row = con.execute("SELECT * FROM cctv_search_requests WHERE site_code = ? AND id = ?", (site_code, request_id)).fetchone()
        con.commit()
    return cctv_request_dict(row)


@app.delete("/api/cctv/requests/{request_id}")
def api_cctv_request_delete(request: Request, request_id: int):
    ensure_ready()
    session = require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    with connect() as con:
        current = require_cctv_request(con, site_code, request_id)
        if not can_delete_cctv_request(session, current):
            raise HTTPException(status_code=403, detail="이 CCTV 요청을 삭제할 권한이 없습니다.")
        con.execute("DELETE FROM cctv_search_requests WHERE site_code = ? AND id = ?", (site_code, request_id))
        con.commit()
    return {"deleted": True, "id": request_id}


@app.get("/api/registry/check", response_model=CheckResponse)
def api_registry_check(request: Request, plate: str):
    require_role(request, VIEW_ROLES)
    return build_check_response(current_site_code(request), plate)


@app.get("/api/registry/search")
def api_registry_search(request: Request, q: str = "", limit: int = 20):
    ensure_ready()
    require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    limit = min(max(limit, 1), 50)
    query = normalize_plate(q) or q.strip()
    like = f"%{query}%"
    with connect() as con:
        rows = con.execute(
            """
            SELECT plate, unit, building, unit_number, owner_name, phone, status, valid_from, valid_to, note, source_file, source_sheet
            FROM vehicles
            WHERE site_code = ?
              AND deleted_at IS NULL
              AND (
                plate LIKE ?
                OR COALESCE(unit, '') LIKE ?
                OR COALESCE(building, '') LIKE ?
                OR COALESCE(unit_number, '') LIKE ?
                OR COALESCE(owner_name, '') LIKE ?
              )
            ORDER BY updated_at DESC, plate
            LIMIT ?
            """,
            (site_code, like, like, like, like, like, limit),
        ).fetchall()
    return [dict(row) for row in rows]


@app.get("/api/registry/status")
def api_registry_status(request: Request):
    ensure_ready()
    require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    source_dir = site_import_dir(site_code)
    with connect() as con:
        vehicle_count = con.execute("SELECT COUNT(*) AS cnt FROM vehicles WHERE site_code = ? AND deleted_at IS NULL", (site_code,)).fetchone()["cnt"]
        manual_count = con.execute("SELECT COUNT(*) AS cnt FROM vehicles WHERE site_code = ? AND deleted_at IS NULL AND COALESCE(manual_override, 0) = 1", (site_code,)).fetchone()["cnt"]
        backups = [
            dict(row)
            for row in con.execute(
                """
                SELECT id, backup_name, vehicles_count, created_by, created_at
                FROM vehicle_backups
                WHERE site_code = ?
                ORDER BY id DESC
                LIMIT 5
                """,
                (site_code,),
            ).fetchall()
        ]
        last_run = con.execute(
            """
            SELECT id, source_dir, files_count, rows_count, imported_at, status, message
            FROM import_runs
            WHERE site_code = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (site_code,),
        ).fetchone()
    return {
        "site_code": site_code,
        "site_name": site_name_for_code(site_code),
        "vehicle_count": vehicle_count,
        "manual_vehicle_count": manual_count,
        "import_dir": str(source_dir),
        "import_files": describe_excel_files(source_dir),
        "backups": backups,
        "ocr_provider": os.getenv("PARKING_OCR_PROVIDER", "tesseract"),
        "ocr_learning": get_learning_status(site_code),
        "last_sync": dict(last_run) if last_run else None,
    }


@app.post("/api/registry/sync")
def api_registry_sync(request: Request, payload: RegistrySyncRequest | None = None):
    ensure_ready()
    require_role(request, {"admin"})
    site_code = current_site_code(request)
    preserve_manual = True if payload is None else payload.preserve_manual
    try:
        with connect() as con:
            backup = create_vehicle_backup(con, site_code, read_session(request).get("u"), f"{site_code}-before-sync-{datetime.now().strftime('%Y%m%d-%H%M%S')}")
            con.commit()
        result = sync_registry_from_dir(site_import_dir(site_code), site_code, preserve_manual=preserve_manual)
        result["backup"] = backup
        return result
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/registry/upload")
async def api_registry_upload(request: Request, files: list[UploadFile] = File(...)):
    ensure_ready()
    require_role(request, {"admin"})
    site_code = current_site_code(request)
    source_dir = site_import_dir(site_code)
    preserve_manual = str(request.query_params.get("preserve_manual", "1")).strip().lower() not in {"0", "false", "no", "off"}

    if not files:
        raise HTTPException(status_code=400, detail="업로드할 Excel 파일을 선택해 주세요.")

    pending: list[tuple[str | None, bytes]] = []
    for item in files:
        payload = await item.read()
        pending.append((item.filename, payload))

    saved_names: set[str] = set()
    uploaded_paths: list[Path] = []
    try:
        for filename, payload in pending:
            try:
                uploaded = store_registry_upload(source_dir, filename, payload, saved_names)
            except ValueError as exc:
                display_name = Path(str(filename or "")).name or "이름 없는 파일"
                raise ValueError(f"{display_name}: {exc}") from exc
            uploaded_paths.append(uploaded)
            saved_names.add(uploaded.name)
        with connect() as con:
            backup = create_vehicle_backup(con, site_code, read_session(request).get("u"), f"{site_code}-before-upload-sync-{datetime.now().strftime('%Y%m%d-%H%M%S')}")
            con.commit()
        sync_result = sync_registry_from_dir(source_dir, site_code, preserve_manual=preserve_manual)
        sync_result["backup"] = backup
    except ValueError as exc:
        for path in uploaded_paths:
            if path.exists():
                path.unlink()
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except FileNotFoundError as exc:
        for path in uploaded_paths:
            if path.exists():
                path.unlink()
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        for path in uploaded_paths:
            if path.exists():
                path.unlink()
        raise HTTPException(status_code=400, detail=f"Excel 파일 처리 중 오류가 발생했습니다: {exc}") from exc

    return {
        "saved_count": len(uploaded_paths),
        "saved_files": [path.name for path in uploaded_paths],
        "import_dir": str(source_dir),
        "sync": sync_result,
    }


@app.get("/api/site/settings")
def api_site_settings(request: Request):
    ensure_ready()
    require_role(request, VIEW_ROLES)
    return site_settings_dict(current_site_code(request))


@app.post("/api/site/settings/capture-placeholder")
async def api_site_settings_capture_placeholder(request: Request, image: UploadFile = File(...)):
    ensure_ready()
    require_role(request, {"admin"})
    site_code = current_site_code(request)
    payload = await image.read()
    image_url = save_site_setting_image(image.filename, payload, site_code)
    with connect() as con:
        con.execute(
            """
            INSERT INTO site_settings(site_code, capture_placeholder_image_url, updated_at)
            VALUES (?, ?, datetime('now'))
            ON CONFLICT(site_code) DO UPDATE SET
              capture_placeholder_image_url = excluded.capture_placeholder_image_url,
              updated_at = datetime('now')
            """,
            (site_code, image_url),
        )
        con.commit()
    return site_settings_dict(site_code)


@app.delete("/api/site/settings/capture-placeholder")
def api_site_settings_capture_placeholder_delete(request: Request):
    ensure_ready()
    require_role(request, {"admin"})
    site_code = current_site_code(request)
    with connect() as con:
        con.execute(
            """
            INSERT INTO site_settings(site_code, capture_placeholder_image_url, updated_at)
            VALUES (?, NULL, datetime('now'))
            ON CONFLICT(site_code) DO UPDATE SET
              capture_placeholder_image_url = NULL,
              updated_at = datetime('now')
            """,
            (site_code,),
        )
        con.commit()
    return site_settings_dict(site_code)


@app.post("/api/ocr/scan")
async def api_ocr_scan(request: Request, photo: UploadFile = File(...), manual_plate: str | None = Form(None)):
    ensure_ready()
    require_role(request, VIEW_ROLES)
    image_bytes = await photo.read()
    if not image_bytes:
        raise HTTPException(status_code=400, detail="사진 파일이 비어 있습니다.")

    scan = scan_plate_image(image_bytes)
    site_code = current_site_code(request)
    best_plate, ordered_candidates = choose_best_scan_candidate(site_code, scan.raw_text, manual_plate, scan.candidates)
    match = build_check_response(current_site_code(request), best_plate).model_dump() if best_plate else None
    return {
        "provider": scan.provider,
        "raw_text": scan.raw_text,
        "candidates": ordered_candidates,
        "best_plate": best_plate,
        "match": match,
        "error": scan.error,
    }


@app.post("/api/enforcement/submit")
async def api_enforcement_submit(
    request: Request,
    plate: str = Form(...),
    inspector: str | None = Form(None),
    location: str | None = Form(None),
    memo: str | None = Form(None),
    raw_ocr_text: str | None = Form(None),
    ocr_best_plate: str | None = Form(None),
    ocr_candidates: str | None = Form(None),
    lat: float | None = Form(None),
    lng: float | None = Form(None),
    photo: UploadFile | None = File(None),
):
    ensure_ready()
    require_role(request, ENFORCEMENT_WRITE_ROLES)
    site_code = current_site_code(request)
    if BILLING_ENFORCEMENT_ENABLED:
        current_usage = billing_status_for_site(site_code)["usage"]["monthly_records"]
        require_billing_capacity(site_code, "monthly_records", int(current_usage) + 1)
    check = build_check_response(site_code, plate)
    photo_path = save_photo(photo, site_code) if photo else None
    learned_candidates = parse_candidates_json(ocr_candidates)

    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO enforcement_events
            (site_code, plate, raw_ocr_text, verdict, verdict_message, unit, owner_name, vehicle_status, inspector, location, memo, photo_path, lat, lng)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                site_code,
                check.plate,
                raw_ocr_text,
                check.verdict,
                check.message,
                check.unit,
                check.owner_name,
                check.status,
                inspector,
                location,
                memo,
                photo_path,
                lat,
                lng,
            ),
        )
        event_id = cur.lastrowid
        row = con.execute("SELECT * FROM enforcement_events WHERE id = ?", (event_id,)).fetchone()
        con.commit()

    record_ocr_feedback(
        site_code=site_code,
        raw_ocr_text=raw_ocr_text,
        suggested_plate=ocr_best_plate,
        corrected_plate=check.plate,
        candidates=learned_candidates,
        photo_path=photo_path,
    )

    return dict(row)


@app.get("/api/enforcement/recent")
def api_enforcement_recent(request: Request, limit: int = 20):
    ensure_ready()
    require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    limit = min(max(limit, 1), 50)
    with connect() as con:
        rows = con.execute(
            """
            SELECT id, plate, verdict, verdict_message, unit, owner_name, inspector, location, memo, photo_path, created_at
            FROM enforcement_events
            WHERE site_code = ?
            ORDER BY id DESC
            LIMIT ?
            """,
            (site_code, limit),
        ).fetchall()
    return [dict(row) for row in rows]


@app.get("/api/registry/vehicles")
def api_registry_vehicles(request: Request, q: str = ""):
    ensure_ready()
    require_vehicle_manager(request)
    site_code = current_site_code(request)
    query = str(q or "").strip()
    if not query:
        return {"items": [], "can_manage": True, "message": "차량번호, 동호수, 차주 또는 연락처를 입력해 조회해 주세요."}

    where = ["site_code = ?", "deleted_at IS NULL"]
    params: list[Any] = [site_code]
    normalized = normalize_plate(query)
    like = f"%{query}%"
    plate_like = f"%{normalized or query}%"
    where.append(
        """
        (
          plate LIKE ?
          OR COALESCE(unit, '') LIKE ?
          OR COALESCE(building, '') LIKE ?
          OR COALESCE(unit_number, '') LIKE ?
          OR COALESCE(owner_name, '') LIKE ?
          OR COALESCE(phone, '') LIKE ?
        )
        """
    )
    params.extend([plate_like, like, like, like, like, like])
    exact_plate = normalized or query
    params.extend([exact_plate, exact_plate])
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT site_code, plate, unit, building, unit_number, owner_name, phone, status, valid_from, valid_to, note, source_file, source_sheet, manual_override, deleted_at, updated_at
            FROM vehicles
            WHERE {' AND '.join(where)}
            ORDER BY
              CASE
                WHEN plate = ? THEN 0
                WHEN plate LIKE ? THEN 1
                ELSE 2
              END,
              manual_override DESC,
              updated_at DESC,
              plate
            LIMIT 1
            """,
            params,
        ).fetchall()
    return {"items": [vehicle_row_dict(row) for row in rows], "can_manage": True}


@app.post("/api/registry/vehicles")
def api_registry_vehicle_create(request: Request, payload: VehicleUpsertRequest):
    ensure_ready()
    session = require_vehicle_manager(request)
    site_code = current_site_code(request)
    values = vehicle_payload_values(payload)
    with connect() as con:
        exists = con.execute("SELECT * FROM vehicles WHERE site_code = ? AND plate = ?", (site_code, values["plate"])).fetchone()
        if exists and not exists["deleted_at"]:
            raise HTTPException(status_code=409, detail="이미 등록된 차량번호입니다.")
        before = vehicle_row_dict(exists)
        con.execute(
            """
            INSERT INTO vehicles(site_code, plate, unit, building, unit_number, owner_name, phone, status, valid_from, valid_to, note, source_file, source_sheet, manual_override, deleted_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'manual', 'manual', 1, NULL, datetime('now'))
            ON CONFLICT(site_code, plate) DO UPDATE SET
              unit = excluded.unit,
              building = excluded.building,
              unit_number = excluded.unit_number,
              owner_name = excluded.owner_name,
              phone = excluded.phone,
              status = excluded.status,
              valid_from = excluded.valid_from,
              valid_to = excluded.valid_to,
              note = excluded.note,
              source_file = 'manual',
              source_sheet = 'manual',
              manual_override = 1,
              deleted_at = NULL,
              updated_at = datetime('now')
            """,
            (
                site_code,
                values["plate"],
                values["unit"],
                values["building"],
                values["unit_number"],
                values["owner_name"],
                values["phone"],
                values["status"],
                values["valid_from"],
                values["valid_to"],
                values["note"],
            ),
        )
        row = con.execute("SELECT * FROM vehicles WHERE site_code = ? AND plate = ?", (site_code, values["plate"])).fetchone()
        after = vehicle_row_dict(row)
        log_vehicle_change(con, site_code, session.get("u"), "create", values["plate"], before, after)
        con.commit()
    return after


@app.patch("/api/registry/vehicles/{plate}")
def api_registry_vehicle_update(request: Request, plate: str, payload: VehicleUpsertRequest):
    ensure_ready()
    session = require_vehicle_manager(request)
    site_code = current_site_code(request)
    current_plate = normalize_plate(plate)
    values = vehicle_payload_values(payload)
    with connect() as con:
        current = con.execute("SELECT * FROM vehicles WHERE site_code = ? AND plate = ? AND deleted_at IS NULL", (site_code, current_plate)).fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="등록차량을 찾을 수 없습니다.")
        before = vehicle_row_dict(current)
        if values["plate"] != current_plate:
            duplicate = con.execute("SELECT 1 FROM vehicles WHERE site_code = ? AND plate = ? AND deleted_at IS NULL", (site_code, values["plate"])).fetchone()
            if duplicate:
                raise HTTPException(status_code=409, detail="변경할 차량번호가 이미 존재합니다.")
        con.execute(
            """
            UPDATE vehicles
            SET plate = ?,
                unit = ?,
                building = ?,
                unit_number = ?,
                owner_name = ?,
                phone = ?,
                status = ?,
                valid_from = ?,
                valid_to = ?,
                note = ?,
                manual_override = 1,
                updated_at = datetime('now')
            WHERE site_code = ? AND plate = ?
            """,
            (
                values["plate"],
                values["unit"],
                values["building"],
                values["unit_number"],
                values["owner_name"],
                values["phone"],
                values["status"],
                values["valid_from"],
                values["valid_to"],
                values["note"],
                site_code,
                current_plate,
            ),
        )
        row = con.execute("SELECT * FROM vehicles WHERE site_code = ? AND plate = ?", (site_code, values["plate"])).fetchone()
        after = vehicle_row_dict(row)
        log_vehicle_change(con, site_code, session.get("u"), "update", values["plate"], before, after)
        con.commit()
    return after


@app.delete("/api/registry/vehicles/{plate}")
def api_registry_vehicle_delete(request: Request, plate: str):
    ensure_ready()
    session = require_vehicle_manager(request)
    site_code = current_site_code(request)
    normalized_plate = normalize_plate(plate)
    with connect() as con:
        current = con.execute("SELECT * FROM vehicles WHERE site_code = ? AND plate = ? AND deleted_at IS NULL", (site_code, normalized_plate)).fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="등록차량을 찾을 수 없습니다.")
        before = vehicle_row_dict(current)
        con.execute(
            "UPDATE vehicles SET deleted_at = datetime('now'), manual_override = 1, updated_at = datetime('now') WHERE site_code = ? AND plate = ?",
            (site_code, normalized_plate),
        )
        log_vehicle_change(con, site_code, session.get("u"), "delete", normalized_plate, before, None)
        con.commit()
    return {"deleted": True, "plate": normalized_plate}


@app.get("/api/registry/backups")
def api_registry_backups(request: Request, limit: int = 20):
    ensure_ready()
    require_vehicle_manager(request)
    site_code = current_site_code(request)
    limit = min(max(limit, 1), 100)
    with connect() as con:
        rows = con.execute(
            """
            SELECT id, backup_name, vehicles_count, created_by, created_at
            FROM vehicle_backups
            WHERE site_code = ?
            ORDER BY id DESC
            LIMIT ?
            """,
            (site_code, limit),
        ).fetchall()
    return [dict(row) for row in rows]


@app.post("/api/registry/backups")
def api_registry_backup_create(request: Request):
    ensure_ready()
    session = require_vehicle_manager(request)
    site_code = current_site_code(request)
    with connect() as con:
        backup = create_vehicle_backup(con, site_code, session.get("u"))
        con.commit()
    return backup


@app.post("/api/registry/backups/{backup_id}/restore")
def api_registry_backup_restore(request: Request, backup_id: int):
    ensure_ready()
    session = require_vehicle_manager(request)
    site_code = current_site_code(request)
    with connect() as con:
        backup_row = con.execute(
            "SELECT * FROM vehicle_backups WHERE site_code = ? AND id = ?",
            (site_code, backup_id),
        ).fetchone()
        if not backup_row:
            raise HTTPException(status_code=404, detail="백업을 찾을 수 없습니다.")
        before_backup = create_vehicle_backup(con, site_code, session.get("u"), f"{site_code}-before-restore-{datetime.now().strftime('%Y%m%d-%H%M%S')}")
        rows = json.loads(backup_row["vehicles_json"] or "[]")
        con.execute("DELETE FROM vehicles WHERE site_code = ?", (site_code,))
        for row in rows:
            con.execute(
                """
                INSERT INTO vehicles(site_code, plate, unit, building, unit_number, owner_name, phone, status, valid_from, valid_to, note, source_file, source_sheet, manual_override, deleted_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, COALESCE(?, datetime('now')))
                """,
                (
                    site_code,
                    row.get("plate"),
                    row.get("unit"),
                    row.get("building"),
                    row.get("unit_number"),
                    row.get("owner_name"),
                    row.get("phone"),
                    row.get("status") or "active",
                    row.get("valid_from"),
                    row.get("valid_to"),
                    row.get("note"),
                    row.get("source_file"),
                    row.get("source_sheet"),
                    1 if row.get("manual_override") else 0,
                    row.get("deleted_at"),
                    row.get("updated_at"),
                ),
            )
        log_vehicle_change(con, site_code, session.get("u"), "restore", None, {"backup_before_restore": before_backup}, {"restored_backup_id": backup_id, "vehicles_count": len(rows)})
        con.commit()
    return {"restored": True, "backup_id": backup_id, "vehicles_count": len(rows), "backup_before_restore": before_backup}


@app.get("/api/enforcement/history")
def api_enforcement_history(
    request: Request,
    q: str = "",
    verdict: str = "",
    date_from: str = "",
    date_to: str = "",
    limit: int = 20,
    offset: int = 0,
):
    ensure_ready()
    require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    limit = min(max(limit, 1), 50)
    offset = max(offset, 0)

    rows = fetch_enforcement_history_rows(
        site_code,
        q=q,
        verdict=verdict,
        date_from=date_from,
        date_to=date_to,
        limit=limit + 1,
        offset=offset,
    )
    has_more = len(rows) > limit
    items = rows[:limit]
    return {
        "items": items,
        "limit": limit,
        "offset": offset,
        "next_offset": offset + len(items) if has_more else None,
        "has_more": has_more,
    }


@app.patch("/api/enforcement/events/{event_id}")
def api_enforcement_event_update(request: Request, event_id: int, payload: EnforcementEventUpdateRequest):
    ensure_ready()
    require_role(request, ENFORCEMENT_WRITE_ROLES)
    site_code = current_site_code(request)
    current = require_enforcement_event(site_code, event_id)

    next_plate = current["plate"]
    if "plate" in payload.model_fields_set:
        normalized_plate = normalize_plate(payload.plate or "")
        if not normalized_plate:
            raise HTTPException(status_code=400, detail="차량번호를 입력해 주세요.")
        next_plate = normalized_plate

    check = build_check_response(site_code, next_plate)
    values = {
        "plate": check.plate,
        "verdict": check.verdict,
        "verdict_message": check.message,
        "unit": check.unit,
        "owner_name": check.owner_name,
        "vehicle_status": check.status,
        "inspector": current.get("inspector"),
        "location": current.get("location"),
        "memo": current.get("memo"),
    }
    if "inspector" in payload.model_fields_set:
        values["inspector"] = str(payload.inspector or "").strip() or None
    if "location" in payload.model_fields_set:
        values["location"] = str(payload.location or "").strip() or None
    if "memo" in payload.model_fields_set:
        values["memo"] = str(payload.memo or "").strip() or None

    with connect() as con:
        con.execute(
            """
            UPDATE enforcement_events
            SET plate = ?,
                verdict = ?,
                verdict_message = ?,
                unit = ?,
                owner_name = ?,
                vehicle_status = ?,
                inspector = ?,
                location = ?,
                memo = ?
            WHERE site_code = ? AND id = ?
            """,
            (
                values["plate"],
                values["verdict"],
                values["verdict_message"],
                values["unit"],
                values["owner_name"],
                values["vehicle_status"],
                values["inspector"],
                values["location"],
                values["memo"],
                site_code,
                event_id,
            ),
        )
        con.commit()
    return require_enforcement_event(site_code, event_id)


@app.delete("/api/enforcement/events/{event_id}")
def api_enforcement_event_delete(request: Request, event_id: int):
    ensure_ready()
    require_role(request, ENFORCEMENT_WRITE_ROLES)
    site_code = current_site_code(request)
    require_enforcement_event(site_code, event_id)
    with connect() as con:
        con.execute("DELETE FROM enforcement_events WHERE site_code = ? AND id = ?", (site_code, event_id))
        con.commit()
    return {"deleted": True, "id": event_id}


@app.get("/api/enforcement/export/rows")
def api_enforcement_export_rows(
    request: Request,
    q: str = "",
    verdict: str = "",
    date_from: str = "",
    date_to: str = "",
    limit: int = 1000,
):
    ensure_ready()
    require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    limit = min(max(limit, 1), 5000)
    rows = fetch_enforcement_history_rows(
        site_code,
        q=q,
        verdict=verdict,
        date_from=date_from,
        date_to=date_to,
        limit=limit + 1,
        offset=0,
    )
    return {
        "site_code": site_code,
        "site_name": site_name_for_code(site_code),
        "items": rows[:limit],
        "limit": limit,
        "truncated": len(rows) > limit,
    }


@app.get("/api/enforcement/export.xlsx")
def api_enforcement_export_xlsx(
    request: Request,
    q: str = "",
    verdict: str = "",
    date_from: str = "",
    date_to: str = "",
):
    ensure_ready()
    require_role(request, VIEW_ROLES)
    site_code = current_site_code(request)
    site_name = site_name_for_code(site_code)
    rows = fetch_enforcement_history_rows(
        site_code,
        q=q,
        verdict=verdict,
        date_from=date_from,
        date_to=date_to,
        limit=5000,
        offset=0,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "불법주차단속대장"
    sheet.merge_cells("A1:L1")
    sheet["A1"] = "불법주차단속대장"
    sheet["A1"].font = Font(size=18, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A2"] = f"아파트: {site_name} ({site_code})"
    sheet["I2"] = f"출력일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    headers = ["장소(층)", "단속시간", "차량번호", "위반내용", "연락처&위치", "경고장", "문자", "통화", "동호수", "차주", "단속자", "판정"]
    header_row = 4
    thin = Side(style="thin", color="8C959F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EAF1EE")

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_index, item in enumerate(rows, start=header_row + 1):
        contact_location = " / ".join(part for part in [item.get("phone"), item.get("location")] if part)
        values = [
            item.get("location") or "",
            str(item.get("created_at") or "").replace("T", " ")[:16],
            item.get("plate") or "",
            item.get("memo") or item.get("verdict_message") or "",
            contact_location,
            "",
            "",
            "",
            item.get("unit") or "",
            item.get("owner_name") or "",
            item.get("inspector") or "",
            item.get("verdict") or "",
        ]
        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(horizontal="center" if col_index in {2, 3, 6, 7, 8, 12} else "left", vertical="center", wrap_text=True)
            cell.border = border

    widths = [18, 18, 16, 28, 28, 10, 10, 10, 16, 16, 14, 12]
    for col_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = width
    sheet.freeze_panes = "A5"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)
    filename = f"불법주차단속대장_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=payload.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
