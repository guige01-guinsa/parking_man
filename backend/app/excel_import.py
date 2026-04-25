from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .db import connect, normalize_site_code
from .plates import normalize_plate, normalize_status

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}


def is_temporary_excel_filename(filename: str | os.PathLike[str] | None) -> bool:
    return Path(str(filename or "")).name.startswith("~$")

FIELD_ALIASES = {
    "plate": {"plate", "carnumber", "platenumber", "차량번호", "번호판", "차번호", "등록번호"},
    "unit": {"unit", "household", "aptunit", "세대", "세대번호", "동호", "동호수", "동호실", "호수"},
    "owner_name": {"owner", "resident", "name", "차주", "입주자", "성명", "이름"},
    "phone": {"phone", "mobile", "tel", "연락처", "전화번호", "휴대폰"},
    "status": {"status", "state", "kind", "상태", "구분", "등록상태"},
    "valid_from": {"validfrom", "startdate", "fromdate", "시작일", "등록시작", "유효시작"},
    "valid_to": {"validto", "enddate", "todate", "만료일", "종료일", "등록종료", "유효종료"},
    "note": {"note", "memo", "remark", "비고", "메모", "특이사항"},
}


@dataclass(slots=True)
class RegistryRecord:
    plate: str
    unit: str | None
    owner_name: str | None
    phone: str | None
    status: str
    valid_from: str | None
    valid_to: str | None
    note: str | None
    source_file: str
    source_sheet: str


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_\-/.()]", "", text)


def resolve_header_field(value: Any) -> str | None:
    token = normalize_header(value)
    if not token:
        return None
    for field, aliases in FIELD_ALIASES.items():
        if token in aliases:
            return field
    return None


def normalize_cell_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def normalize_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(".", "-").replace("/", "-")
    try:
        return date.fromisoformat(text[:10]).isoformat()
    except ValueError:
        return text[:10]


def find_header_row(worksheet) -> tuple[int, dict[str, int]] | None:
    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        mapping: dict[str, int] = {}
        for index, value in enumerate(row):
            field = resolve_header_field(value)
            if field:
                mapping[field] = index
        if "plate" in mapping:
            return row_index, mapping
    return None


def build_record(row: tuple[Any, ...], mapping: dict[str, int], source_file: str, source_sheet: str) -> RegistryRecord | None:
    plate = normalize_plate(row[mapping["plate"]] if mapping["plate"] < len(row) else None)
    if not plate:
        return None
    return RegistryRecord(
        plate=plate,
        unit=normalize_cell_text(row[mapping["unit"]]) if "unit" in mapping and mapping["unit"] < len(row) else None,
        owner_name=normalize_cell_text(row[mapping["owner_name"]]) if "owner_name" in mapping and mapping["owner_name"] < len(row) else None,
        phone=normalize_cell_text(row[mapping["phone"]]) if "phone" in mapping and mapping["phone"] < len(row) else None,
        status=normalize_status(row[mapping["status"]]) if "status" in mapping and mapping["status"] < len(row) else "active",
        valid_from=normalize_date(row[mapping["valid_from"]]) if "valid_from" in mapping and mapping["valid_from"] < len(row) else None,
        valid_to=normalize_date(row[mapping["valid_to"]]) if "valid_to" in mapping and mapping["valid_to"] < len(row) else None,
        note=normalize_cell_text(row[mapping["note"]]) if "note" in mapping and mapping["note"] < len(row) else None,
        source_file=source_file,
        source_sheet=source_sheet,
    )


def list_excel_files(source_dir: Path) -> list[Path]:
    return sorted(
        [
            path
            for path in source_dir.iterdir()
            if path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES and not is_temporary_excel_filename(path.name)
        ],
        key=lambda item: item.name.lower(),
    )


def build_safe_excel_filename(filename: str | None, existing_names: set[str] | None = None) -> str:
    raw_name = Path(str(filename or "")).name.strip()
    if not raw_name:
        raise ValueError("업로드 파일 이름이 비어 있습니다.")
    if is_temporary_excel_filename(raw_name):
        raise ValueError("Excel 임시 잠금 파일은 업로드할 수 없습니다. 원본 파일을 선택해 주세요.")

    stem = re.sub(r"[^\w가-힣.-]+", "-", Path(raw_name).stem).strip("-.")
    suffix = Path(raw_name).suffix.lower()
    if suffix not in EXCEL_SUFFIXES:
        raise ValueError("Excel 파일은 .xlsx 또는 .xlsm 형식만 업로드할 수 있습니다.")
    if not stem:
        stem = "registry"

    candidate = f"{stem}{suffix}"
    existing_names = existing_names or set()
    counter = 2
    while candidate in existing_names:
        candidate = f"{stem}-{counter}{suffix}"
        counter += 1
    return candidate


def store_registry_upload(source_dir: str | os.PathLike[str], filename: str | None, data: bytes, existing_names: set[str] | None = None) -> Path:
    if not data:
        raise ValueError("비어 있는 파일은 업로드할 수 없습니다.")

    source_path = Path(source_dir)
    source_path.mkdir(parents=True, exist_ok=True)
    reserved_names = {path.name for path in list_excel_files(source_path)}
    reserved_names.update(existing_names or set())

    safe_name = build_safe_excel_filename(filename, reserved_names)
    target = source_path / safe_name
    target.write_bytes(data)
    return target


def describe_excel_files(source_dir: str | os.PathLike[str], limit: int = 20) -> list[dict[str, Any]]:
    source_path = Path(source_dir)
    if not source_path.exists():
        return []

    files = sorted(list_excel_files(source_path), key=lambda item: item.stat().st_mtime, reverse=True)
    rows: list[dict[str, Any]] = []
    for path in files[: max(limit, 0)]:
        stat = path.stat()
        rows.append(
            {
                "name": path.name,
                "size": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
            }
        )
    return rows


def record_import_run(site_code: str, source_dir: Path, files_count: int, rows_count: int, status: str, message: str) -> None:
    with connect() as con:
        con.execute(
            """
            INSERT INTO import_runs(site_code, source_dir, files_count, rows_count, status, message)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (site_code, str(source_dir), files_count, rows_count, status, message),
        )
        con.commit()


def sync_registry_from_dir(source_dir: str | os.PathLike[str], site_code: str | None = None) -> dict[str, Any]:
    source_path = Path(source_dir)
    source_path.mkdir(parents=True, exist_ok=True)
    resolved_site_code = normalize_site_code(site_code)
    files = list_excel_files(source_path)

    if not files:
        message = f"Excel 파일이 없습니다. 폴더: {source_path}"
        record_import_run(resolved_site_code, source_path, 0, 0, "skipped", message)
        raise FileNotFoundError(message)

    merged: dict[str, RegistryRecord] = {}
    rows_seen = 0

    for excel_file in files:
        workbook = load_workbook(excel_file, data_only=True, read_only=True)
        for worksheet in workbook.worksheets:
            header = find_header_row(worksheet)
            if not header:
                continue
            header_row, mapping = header
            for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                record = build_record(row, mapping, excel_file.name, worksheet.title)
                if not record:
                    continue
                rows_seen += 1
                merged[record.plate] = record
        workbook.close()

    if not merged:
        message = "유효한 차량번호 행을 찾지 못했습니다."
        record_import_run(resolved_site_code, source_path, len(files), 0, "failed", message)
        raise ValueError(message)

    with connect() as con:
        con.execute("DELETE FROM vehicles WHERE site_code = ?", (resolved_site_code,))
        con.executemany(
            """
            INSERT INTO vehicles
            (site_code, plate, unit, owner_name, phone, status, valid_from, valid_to, note, source_file, source_sheet, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            [
                (
                    resolved_site_code,
                    record.plate,
                    record.unit,
                    record.owner_name,
                    record.phone,
                    record.status,
                    record.valid_from,
                    record.valid_to,
                    record.note,
                    record.source_file,
                    record.source_sheet,
                )
                for record in merged.values()
            ],
        )
        con.execute(
            """
            INSERT INTO import_runs(site_code, source_dir, files_count, rows_count, status, message)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                resolved_site_code,
                str(source_path),
                len(files),
                len(merged),
                "success",
                f"{len(files)}개 파일, {rows_seen}개 행 처리, {len(merged)}대 반영",
            ),
        )
        con.commit()

    return {
        "site_code": resolved_site_code,
        "source_dir": str(source_path),
        "files_count": len(files),
        "rows_seen": rows_seen,
        "vehicles_loaded": len(merged),
    }

